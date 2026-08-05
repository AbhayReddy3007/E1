#!/usr/bin/env python3
"""
fetcher.py – Post-process enrichment + clinical efficacy scoring for main1.py Excel output.

Step 1 – Gemini enrichment (parallel):
  Reads the Excel file produced by main1.py and uses Gemini (with Google Search)
  to fill in the following per-trial columns:

    dosage
    hba1c_change_pct  hba1c_duration  hba1c_rationale  hba1c_confidence
    weight_change_pct weight_duration weight_rationale  weight_confidence
    alt_reduction_pct alt_duration    alt_rationale     alt_confidence
    mash_change_pct   mash_duration   mash_rationale    mash_confidence

Step 2 – Dimension III Clinical Efficacy Scoring:
  After enrichment, scores the molecule across all four endpoints using a
  phase-anchored algorithm and writes a "Score Summary" sheet to the workbook:

    Scoring rules:
      Phase 3 -> no penalty  |  Phase 2 -> x0.85  |  Phase 1 -> x0.65
      >=22% -> 5  |  16-21.9% -> 4  |  10-15.9% -> 3  |  5-9.9% -> 2  |  <5% -> 1
    Weights: Weight Loss 40% | HbA1c 40% | MASH 10% | ALT 10%

  A Gemini-generated narrative rationale is also produced for the final score.

Usage:
    python fetcher.py <molecule_name> [--excel path.xlsx] [--workers N] [--out enriched.xlsx]

    # Typical: molecule name only - auto-discovers <molecule>_trials.xlsx
    python fetcher.py Cagrisema

    # Explicit paths / parallelism
    python fetcher.py Cagrisema --excel cagrisema_trials.xlsx --workers 8 --out cagrisema_enriched.xlsx

    # Skip scoring step
    python fetcher.py Cagrisema --no-score
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
import time
from typing import Any, Dict, List, Optional

# -- third-party ---------------------------------------------------------------
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl not installed.  Run: pip install openpyxl")

try:
    from google import genai
    from google.genai import types
except ImportError:
    sys.exit("ERROR: google-genai not installed.  Run: pip install google-genai")

try:
    from json_repair import repair_json
    _HAS_JSON_REPAIR = True
except ImportError:
    _HAS_JSON_REPAIR = False

# -- load .env -----------------------------------------------------------------
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ==============================================================================
# SECTION 1 – CONSTANTS & COLUMN DEFINITIONS
# ==============================================================================

MODEL           = "gemini-2.5-flash"
RATIONALE_MODEL = "gemini-2.5-flash"   # used for the narrative score rationale
MAX_RETRIES     = 5
INITIAL_BACKOFF = 2.0
BATCH_SIZE      = 6
DEFAULT_WORKERS = 6

# Columns fetcher fills per trial
OUTCOME_COLS = [
    "dosage",
    "hba1c_change_pct",  "hba1c_duration",  "hba1c_rationale",  "hba1c_confidence",
    "weight_change_pct", "weight_duration",  "weight_rationale", "weight_confidence",
    "alt_reduction_pct", "alt_duration",     "alt_rationale",    "alt_confidence",
    "mash_change_pct",   "mash_duration",    "mash_rationale",   "mash_confidence",
]

# Full ordered column list matching main1.py
ALL_COLUMNS = [
    "molecule_name", "registry_source", "trial_id", "acronym",
    "dosage", "phase", "trial_title", "trial_study", "trial_size",
    "trial_location", "trial_start_date", "trial_completion_date", "phase_status",
    "hba1c_change_pct",  "hba1c_duration",  "hba1c_rationale",  "hba1c_confidence",
    "weight_change_pct", "weight_duration",  "weight_rationale", "weight_confidence",
    "alt_reduction_pct", "alt_duration",     "alt_rationale",    "alt_confidence",
    "mash_change_pct",   "mash_duration",    "mash_rationale",   "mash_confidence",
    "company_name", "source_url",
]

COLUMN_WIDTHS = {
    "molecule_name": 16, "registry_source": 18, "trial_id": 18,
    "acronym": 18, "dosage": 18, "phase": 10, "trial_title": 40,
    "trial_study": 26, "trial_size": 12, "trial_location": 24,
    "trial_start_date": 16, "trial_completion_date": 18, "phase_status": 18,
    "hba1c_change_pct": 16,  "hba1c_duration": 14,  "hba1c_rationale": 45,  "hba1c_confidence": 14,
    "weight_change_pct": 16, "weight_duration": 14,  "weight_rationale": 45, "weight_confidence": 14,
    "alt_reduction_pct": 16, "alt_duration": 14,     "alt_rationale": 45,    "alt_confidence": 14,
    "mash_change_pct": 18,   "mash_duration": 14,    "mash_rationale": 45,   "mash_confidence": 14,
    "company_name": 28, "source_url": 40,
}

# ==============================================================================
# SECTION 2 – GEMINI CLIENT
# ==============================================================================

class _NoApiKeyError(RuntimeError):
    """Raised instead of sys.exit so async threads propagate cleanly."""


def _make_client() -> genai.Client:
    api_key = os.getenv("GOOGLE_API_KEY") or os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise _NoApiKeyError(
            "No API key found.\n"
            "  Option 1: add  GEMINI_API_KEY=your_key  to your .env file\n"
            "  Option 2: set the environment variable before running:\n"
            "            Windows:  set GEMINI_API_KEY=your_key\n"
            "            Mac/Linux: export GEMINI_API_KEY=your_key"
        )
    return genai.Client(api_key=api_key)


_client: Optional[genai.Client] = None


def get_client() -> genai.Client:
    global _client
    if _client is None:
        _client = _make_client()
    return _client


def _check_api_key_early() -> None:
    """Validate key before launching async – gives a clean error message."""
    try:
        get_client()
    except _NoApiKeyError as exc:
        print(f"\nERROR: {exc}\n", file=sys.stderr)
        sys.exit(1)


# ==============================================================================
# SECTION 3 – JSON HELPERS
# ==============================================================================

def _safe_parse(text: str) -> Any:
    """Parse JSON from a Gemini response, with repair fallback."""
    if not text:
        return None
    text = text.strip()
    if "```json" in text:
        text = text.split("```json", 1)[1].split("```", 1)[0].strip()
    elif "```" in text:
        parts = text.split("```")
        if len(parts) >= 3:
            text = parts[1].strip()
    for i, ch in enumerate(text):
        if ch in "{[":
            text = text[i:]
            break
    if _HAS_JSON_REPAIR:
        try:
            return repair_json(text, return_objects=True)
        except Exception:
            pass
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    return None


# ==============================================================================
# SECTION 4 – GEMINI CALL (ASYNC, WITH RETRY)
# ==============================================================================

def _sync_call(prompt: str, use_search: bool = True) -> str:
    """Blocking Gemini call, optionally with Google Search grounding."""
    client = get_client()
    contents = [types.Content(role="user", parts=[types.Part.from_text(text=prompt)])]
    config_kwargs: Dict[str, Any] = {}
    if use_search:
        config_kwargs["tools"] = [types.Tool(googleSearch=types.GoogleSearch())]
    config = types.GenerateContentConfig(**config_kwargs)
    out = ""
    for chunk in client.models.generate_content_stream(
        model=MODEL, contents=contents, config=config
    ):
        if chunk.text:
            out += chunk.text
    return out.strip()


async def _gemini_call(prompt: str, use_search: bool = True) -> str:
    """Async Gemini call with exponential backoff on rate-limit errors."""
    backoff = INITIAL_BACKOFF
    for attempt in range(MAX_RETRIES + 1):
        try:
            return await asyncio.to_thread(_sync_call, prompt, use_search)
        except _NoApiKeyError:
            raise
        except Exception as exc:
            err = str(exc).lower()
            if any(k in err for k in ("429", "rate limit", "quota", "resource exhausted")):
                if attempt == MAX_RETRIES:
                    print(f"  X Max retries exceeded: {exc}", file=sys.stderr)
                    raise
                print(
                    f"  ! Rate-limit – waiting {backoff:.0f}s "
                    f"(attempt {attempt+1}/{MAX_RETRIES})...",
                    file=sys.stderr,
                )
                await asyncio.sleep(backoff)
                backoff *= 2
            else:
                print(f"  X Gemini error: {exc}", file=sys.stderr)
                raise
    return ""


# ==============================================================================
# SECTION 5 – ENRICHMENT (PER-TRIAL EFFICACY FETCH)
# ==============================================================================

def _build_prompt(molecule: str, batch: List[Dict[str, str]]) -> str:
    trial_lines = "\n".join(
        f"  - {t.get('trial_id','?')} | {t.get('trial_title','')[:80]} "
        f"| Phase {t.get('phase','?')} | {t.get('company_name','')} "
        f"| URL: {t.get('source_url','')}"
        for t in batch
    )
    return f"""You are a clinical data extraction engine with access to Google Search and live trial registries.

MOLECULE: {molecule}

TRIALS TO ENRICH ({len(batch)} total):
{trial_lines}

For EACH trial listed above, search ClinicalTrials.gov, PubMed, and published results to extract:

1. dosage          - Primary or highest dose tested (e.g. "2.4 mg OW", "15 mg QD").
                     If multiple doses, pick the highest. Format: "[amount] [unit] [frequency]"

2. hba1c_change_pct  - HbA1c reduction in percentage points (positive number, e.g. "1.8").
                        "N/A" if not a diabetes trial or data unavailable.
   hba1c_duration     - Timepoint of measurement (e.g. "26 wk"). "N/A" if unavailable.
   hba1c_rationale    - 1-2 sentences: state the exact source (paper / registry / result date)
                        and why this value was chosen (e.g. primary endpoint, highest-dose arm).
   hba1c_confidence   - "High" / "Medium" / "Low" reflecting data reliability.

3. weight_change_pct - Body weight loss percentage (positive number, e.g. "15.2"). "N/A" if unavailable.
   weight_duration    - Timepoint (e.g. "68 wk"). "N/A" if unavailable.
   weight_rationale   - 1-2 sentences citing source and reason for value chosen.
   weight_confidence  - "High" / "Medium" / "Low".

4. alt_reduction_pct - ALT enzyme reduction percentage (positive number). "N/A" if unavailable.
   alt_duration       - Timepoint (e.g. "24 wk"). "N/A" if unavailable.
   alt_rationale      - 1-2 sentences citing source and reason.
   alt_confidence     - "High" / "Medium" / "Low".

5. mash_change_pct   - MASH/NASH resolution rate or fibrosis improvement % (positive number).
                        "N/A" if not a liver trial.
   mash_duration      - Timepoint (e.g. "72 wk"). "N/A" if unavailable.
   mash_rationale     - 1-2 sentences citing source and reason.
   mash_confidence    - "High" / "Medium" / "Low".

RULES:
- Use actual published results where available; fall back to registry data.
- Report reductions as POSITIVE numbers.
- Use "N/A" for fields with genuinely no data.
- Each rationale MUST name the specific source (e.g. "NEJM 2023 STEP 1 paper", "ClinicalTrials.gov NCT03548935 results section").
- One JSON object per trial keyed by trial_id.

Return ONLY valid JSON, no markdown, no preamble:

{{
  "results": {{
    "<trial_id_1>": {{
      "dosage": "...",
      "hba1c_change_pct": "...", "hba1c_duration": "...", "hba1c_rationale": "...", "hba1c_confidence": "...",
      "weight_change_pct": "...", "weight_duration": "...", "weight_rationale": "...", "weight_confidence": "...",
      "alt_reduction_pct": "...", "alt_duration": "...", "alt_rationale": "...", "alt_confidence": "...",
      "mash_change_pct": "...", "mash_duration": "...", "mash_rationale": "...", "mash_confidence": "..."
    }},
    "<trial_id_2>": {{ ... }}
  }}
}}
"""


async def _enrich_batch(
    molecule: str,
    batch: List[Dict[str, str]],
    batch_idx: int,
    total_batches: int,
    semaphore: asyncio.Semaphore,
) -> Dict[str, Dict[str, str]]:
    """Enrich one batch of trials; returns {trial_id: {field: value}}."""
    async with semaphore:
        ids = [t.get("trial_id", "?") for t in batch]
        print(f"  Batch {batch_idx+1}/{total_batches} -> {ids}", file=sys.stderr)
        try:
            raw = await _gemini_call(_build_prompt(molecule, batch), use_search=True)
        except Exception:
            return {}
        data = _safe_parse(raw)
        if not data:
            print(f"  X Batch {batch_idx+1}: could not parse response", file=sys.stderr)
            return {}
        if isinstance(data, dict) and "results" in data:
            results = data["results"]
        elif isinstance(data, dict):
            results = data
        else:
            print(f"  X Batch {batch_idx+1}: unexpected JSON structure", file=sys.stderr)
            return {}
        print(f"  OK Batch {batch_idx+1}: enriched {len(results)} trial(s)", file=sys.stderr)
        return results


async def enrich_all(
    molecule: str,
    rows: List[Dict[str, str]],
    max_workers: int = DEFAULT_WORKERS,
) -> List[Dict[str, str]]:
    """Enrich all rows in parallel batches."""
    batches = [rows[i: i + BATCH_SIZE] for i in range(0, len(rows), BATCH_SIZE)]
    total = len(batches)
    print(
        f"\n[ENRICH] {len(rows)} trial(s) across {total} batch(es) "
        f"(max {max_workers} concurrent)...\n",
        file=sys.stderr,
    )
    semaphore = asyncio.Semaphore(max_workers)

    async def _staggered(coro, delay: float):
        await asyncio.sleep(delay)
        return await coro

    staggered_tasks = [
        _staggered(_enrich_batch(molecule, batch, idx, total, semaphore), idx * 0.4)
        for idx, batch in enumerate(batches)
    ]
    batch_results = await asyncio.gather(*staggered_tasks, return_exceptions=False)

    # Merge all results keyed by trial_id
    merged: Dict[str, Dict[str, str]] = {}
    for br in batch_results:
        if isinstance(br, dict):
            merged.update(br)

    # Apply back to rows
    updated = 0
    for row in rows:
        tid = row.get("trial_id", "")
        enrichment = merged.get(tid, {})
        if not enrichment:
            for k, v in merged.items():
                if k.strip().upper() == tid.strip().upper():
                    enrichment = v
                    break
        if enrichment:
            for col in OUTCOME_COLS:
                val = enrichment.get(col, "")
                if val and str(val).strip().lower() not in ("n/a", "null", "none", ""):
                    row[col] = str(val).strip()
                elif col not in row or not row[col]:
                    row[col] = "N/A"
            updated += 1
        else:
            for col in OUTCOME_COLS:
                if col not in row or not row[col]:
                    row[col] = "N/A"

    print(f"\n[ENRICH] Done: {updated}/{len(rows)} trial(s) updated.\n", file=sys.stderr)
    return rows


# ==============================================================================
# SECTION 6 – DIMENSION III CLINICAL EFFICACY SCORING
# (Ported directly from clinical_efficacy_scorer.py)
# ==============================================================================

SCORE_TABLE = [
    (22.0, 5),
    (16.0, 4),
    (10.0, 3),
    (5.0,  2),
    (0.0,  1),
]

ENDPOINT_WEIGHTS = {
    "weight_loss": 0.40,
    "hba1c":       0.40,
    "mash":        0.10,
    "alt":         0.10,
}

# Maps endpoint key -> field name in the trial dict
FIELD_MAP = {
    "weight_loss": "weight_change_pct",
    "hba1c":       "hba1c_change_pct",
    "mash":        "mash_change_pct",     # main1.py uses mash_change_pct
    "alt":         "alt_reduction_pct",
}

PHASE_PENALTY = {3: 1.00, 2: 0.85, 1: 0.65}


def _parse_phase(raw) -> Optional[int]:
    """Normalise phase string/number to int 1/2/3."""
    if raw is None:
        return None
    s = str(raw).strip().upper().replace("PHASE", "").strip()
    if s.startswith("3"):
        return 3
    if s.startswith("2"):
        return 2
    if s.startswith("1"):
        return 1
    try:
        v = float(s)
        if v >= 3:
            return 3
        if v >= 2:
            return 2
        return 1
    except ValueError:
        return None


def _parse_float(raw) -> Optional[float]:
    """Return float or None for missing/N/A values."""
    if raw is None:
        return None
    s = str(raw).strip()
    if s.lower() in ("n/a", "", "0", "none", "null"):
        return None
    # Strip trailing % signs if present
    s = s.rstrip("%").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _pct_to_score(pct: float) -> int:
    """Convert adjusted % to 1-5 score using the documented threshold table."""
    for threshold, score in SCORE_TABLE:
        if pct >= threshold:
            return score
    return 1


def _score_endpoint(trials: List[Dict[str, str]], value_field: str) -> Dict[str, Any]:
    """
    Score a single endpoint across all trials.

    Algorithm (from scorer docs):
      1. Parse & validate all trials
      2. Phase-anchored selection: prefer Phase 3, fall back to Phase 2 (x0.85), Phase 1 (x0.65)
      3. Within selected phase: deduplicate by trial_id, keep highest dosage per trial
      4. Take highest value across all deduplicated trials in that phase
      5. Apply phase penalty -> convert to 1-5 score
    """
    valid = []
    for t in trials:
        phase   = _parse_phase(t.get("phase"))
        value   = _parse_float(t.get(value_field))
        n       = _parse_float(t.get("trial_size")) or 0
        trial_id = t.get("trial_id") or t.get("Trial ID")

        if phase is None or value is None or n <= 0:
            continue
        if not trial_id:
            trial_id = f"__unknown_{id(t)}"

        valid.append({
            "phase": phase, "value": value, "n": n,
            "trial_id": trial_id, "full_trial": t,
        })

    if not valid:
        return {
            "best_value": None, "raw_value": None,
            "phase_used": None, "penalty": 1.0,
            "score": None, "trial_details": {},
            "reason": "No valid data for this endpoint",
        }

    for target_phase in (3, 2, 1):
        phase_trials = [r for r in valid if r["phase"] == target_phase]
        if not phase_trials:
            continue

        # Deduplicate: keep highest-performing dosage per trial
        trial_groups: Dict[str, list] = {}
        for t in phase_trials:
            trial_groups.setdefault(t["trial_id"], []).append(t)
        deduplicated = [max(arms, key=lambda x: x["value"]) for arms in trial_groups.values()]

        best = max(deduplicated, key=lambda r: r["value"])
        raw  = best["value"]
        pen  = PHASE_PENALTY[target_phase]
        adj  = raw * pen

        ft = best.get("full_trial", {})
        trial_details = {
            "trial_id":       best["trial_id"],
            "dosage":         ft.get("dosage", "N/A"),
            "weight_duration": ft.get("weight_duration", "N/A"),
            "hba1c_duration":  ft.get("hba1c_duration", "N/A"),
            "mash_duration":   ft.get("mash_duration", "N/A"),
            "alt_duration":    ft.get("alt_duration", "N/A"),
        }

        return {
            "best_value":   round(adj, 4),
            "raw_value":    round(raw, 4),
            "phase_used":   target_phase,
            "penalty":      pen,
            "score":        _pct_to_score(adj),
            "trial_details": trial_details,
            "reason": (
                f"Phase {target_phase} data used"
                + (f" (x{pen} penalty applied)" if pen < 1 else "")
            ),
        }

    return {
        "best_value": None, "raw_value": None,
        "phase_used": None, "penalty": 1.0,
        "score": None, "trial_details": {},
        "reason": "Unexpected state",
    }


def compute_clinical_efficacy_score(molecule: str, rows: List[Dict[str, str]]) -> Dict[str, Any]:
    """
    Compute the Dimension III clinical efficacy score (1-5 scale) for a molecule.

    Args:
        molecule: Drug name
        rows:     List of enriched trial dicts (from enrich_all)

    Returns dict with keys: molecule, total_trials, endpoints, weighted_score,
                            score_breakdown, data_coverage
    """
    total = len(rows)
    endpoint_results = {ep: _score_endpoint(rows, field) for ep, field in FIELD_MAP.items()}

    score_sum   = 0.0
    scored_eps  = []
    missing_eps = []
    for ep, result in endpoint_results.items():
        w = ENDPOINT_WEIGHTS[ep]
        if result["score"] is not None:
            score_sum += result["score"] * w
            scored_eps.append(ep)
        else:
            missing_eps.append(ep)

    weighted_score = score_sum

    lines = []
    for ep, result in endpoint_results.items():
        w_pct = int(ENDPOINT_WEIGHTS[ep] * 100)
        if result["score"] is not None:
            lines.append(
                f"  {ep:12} | adj={result['best_value']:.2f}%  "
                f"score={result['score']}  weight={w_pct}%  ({result['reason']})"
            )
        else:
            lines.append(f"  {ep:12} | N/A  weight={w_pct}%  ({result['reason']})")

    coverage = (
        f"{len(scored_eps)}/4 endpoints scored"
        + (f" (missing: {', '.join(missing_eps)})" if missing_eps else "")
    )

    return {
        "molecule":        molecule,
        "total_trials":    total,
        "endpoints":       endpoint_results,
        "weighted_score":  round(weighted_score, 3),
        "score_breakdown": "\n".join(lines),
        "data_coverage":   coverage,
    }


# ==============================================================================
# SECTION 7 – NARRATIVE RATIONALE GENERATION (via Gemini, no search needed)
# ==============================================================================

def generate_score_rationale(molecule: str, score_result: Dict[str, Any]) -> str:
    """
    Call Gemini to produce a 4-5 paragraph clinical narrative explaining the score.
    Runs synchronously (called once after all async work is done).
    """
    print("\n[SCORE] Generating narrative rationale via Gemini...", file=sys.stderr)
    endpoints = score_result.get("endpoints", {})

    endpoint_summary = {
        "Weight Loss (40% weight)": {
            "score":      endpoints.get("weight_loss", {}).get("score"),
            "best_value": endpoints.get("weight_loss", {}).get("best_value"),
            "raw_value":  endpoints.get("weight_loss", {}).get("raw_value"),
            "phase_used": endpoints.get("weight_loss", {}).get("phase_used"),
            "trial_used": endpoints.get("weight_loss", {}).get("trial_details", {}).get("trial_id"),
            "dosage":     endpoints.get("weight_loss", {}).get("trial_details", {}).get("dosage"),
            "duration":   endpoints.get("weight_loss", {}).get("trial_details", {}).get("weight_duration"),
        },
        "HbA1c Reduction (40% weight)": {
            "score":      endpoints.get("hba1c", {}).get("score"),
            "best_value": endpoints.get("hba1c", {}).get("best_value"),
            "raw_value":  endpoints.get("hba1c", {}).get("raw_value"),
            "phase_used": endpoints.get("hba1c", {}).get("phase_used"),
            "trial_used": endpoints.get("hba1c", {}).get("trial_details", {}).get("trial_id"),
            "dosage":     endpoints.get("hba1c", {}).get("trial_details", {}).get("dosage"),
            "duration":   endpoints.get("hba1c", {}).get("trial_details", {}).get("hba1c_duration"),
        },
        "MASH Resolution (10% weight)": {
            "score":      endpoints.get("mash", {}).get("score"),
            "best_value": endpoints.get("mash", {}).get("best_value"),
            "raw_value":  endpoints.get("mash", {}).get("raw_value"),
            "phase_used": endpoints.get("mash", {}).get("phase_used"),
            "trial_used": endpoints.get("mash", {}).get("trial_details", {}).get("trial_id"),
            "dosage":     endpoints.get("mash", {}).get("trial_details", {}).get("dosage"),
            "duration":   endpoints.get("mash", {}).get("trial_details", {}).get("mash_duration"),
        },
        "ALT Reduction (10% weight)": {
            "score":      endpoints.get("alt", {}).get("score"),
            "best_value": endpoints.get("alt", {}).get("best_value"),
            "raw_value":  endpoints.get("alt", {}).get("raw_value"),
            "phase_used": endpoints.get("alt", {}).get("phase_used"),
            "trial_used": endpoints.get("alt", {}).get("trial_details", {}).get("trial_id"),
            "dosage":     endpoints.get("alt", {}).get("trial_details", {}).get("dosage"),
            "duration":   endpoints.get("alt", {}).get("trial_details", {}).get("alt_duration"),
        },
    }

    prompt = f"""You are a clinical pharmacology expert. Generate a comprehensive, evidence-based rationale explaining the clinical efficacy score for {molecule}.

SCORING RESULTS:
- Clinical Efficacy Score: {score_result['weighted_score']} / 5
- Coverage: {score_result['data_coverage']}

ENDPOINT PERFORMANCE (these are the EXACT trials used for scoring):
{json.dumps(endpoint_summary, indent=2)}

SCORING METHODOLOGY:
- Score ranges: 5 = >=22%, 4 = 16-21.9%, 3 = 10-15.9%, 2 = 5-9.9%, 1 = <5%
- Phase penalties: Phase 3 = no penalty, Phase 2 = x0.85, Phase 1 = x0.65
- Weighted average: Weight Loss (40%) + HbA1c (40%) + MASH (10%) + ALT (10%)

YOUR TASK:
Write a comprehensive 4-5 paragraph clinical rationale that:

1. Opening paragraph: State the clinical efficacy score and provide a high-level summary of {molecule}'s clinical performance.

2. Weight Loss paragraph: Discuss weight loss efficacy - MUST reference specific percentage, duration, dosage, phase, trial ID, and comparison to clinical benchmarks (>15% excellent, 10-15% good).

3. HbA1c Reduction paragraph: Discuss glycemic control - MUST reference specific percentage, duration, dosage, phase, trial ID (>2% excellent, 1-2% good for diabetes management).

4. MASH/ALT paragraph (if data available): Discuss liver endpoints with duration context. If N/A, briefly explain why.

5. Concluding paragraph: Summarize score rationale, data quality (phase distribution, sample sizes), and why this score represents the molecule's true clinical efficacy profile.

WRITING GUIDELINES:
- Include dosage and duration for each endpoint (e.g., "achieved 15% weight loss at 2.4 mg over 68 weeks")
- Use specific numbers from the data (percentages, trial IDs, phase info)
- Be objective and evidence-based with medical terminology
- Plain text with paragraph breaks only - no markdown, no headers, no bullets
- Do not use phrases like "The rationale is..." - write the rationale directly
- Write as documentation for regulatory or pharma stakeholders

IMPORTANT: Use trial_id, dosage, and duration ONLY from the ENDPOINT PERFORMANCE section above. Do not invent or infer from other sources.

Generate the rationale now:"""

    try:
        client = get_client()
        contents = [types.Content(role="user", parts=[types.Part.from_text(text=prompt)])]
        config = types.GenerateContentConfig(
            temperature=0.3,
            response_mime_type="text/plain",
        )
        response_text = ""
        for chunk in client.models.generate_content_stream(
            model=RATIONALE_MODEL, contents=contents, config=config
        ):
            if chunk.text:
                response_text += chunk.text
        rationale = response_text.strip().replace("\n\n\n", "\n\n")
        print("[SCORE] Rationale generated.", file=sys.stderr)
        return rationale
    except Exception as exc:
        print(f"  [SCORE] Rationale generation failed: {exc}", file=sys.stderr)
        return (
            f"{molecule} received a clinical efficacy score of "
            f"{score_result['weighted_score']}/5 based on analysis of "
            f"{score_result['total_trials']} trials. {score_result['data_coverage']}."
        )


# ==============================================================================
# SECTION 8 – EXCEL I/O
# ==============================================================================

def _read_excel(path: str) -> List[Dict[str, str]]:
    """Load rows from main1.py Excel output into a list of dicts."""
    wb = load_workbook(path)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows: List[Dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record: Dict[str, str] = {h: (str(v) if v is not None else "") for h, v in zip(headers, row)}
        if any(record.values()):
            rows.append(record)
    print(f"[INPUT] Loaded {len(rows)} row(s) from {path}", file=sys.stderr)
    return rows


def _write_excel(
    rows: List[Dict[str, str]],
    path: str,
    score_result: Optional[Dict[str, Any]] = None,
    score_rationale: Optional[str] = None,
) -> None:
    """Write enriched rows + optional Score Summary sheet to the workbook."""
    from openpyxl import Workbook

    wb = Workbook()

    # ---- Sheet 1: Clinical Trials (Enriched) ---------------------------------
    ws = wb.active
    ws.title = "Clinical Trials (Enriched)"

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    DATA_FONT   = Font(name="Arial", size=9)
    WRAP_ALIGN  = Alignment(wrap_text=True, vertical="top")
    THIN        = Side(style="thin", color="D0D0D0")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ALT_FILL    = PatternFill("solid", fgColor="EBF5FB")
    ENRICH_FILL = PatternFill("solid", fgColor="E8F8E8")

    extra_cols = [c for c in (rows[0].keys() if rows else []) if c not in ALL_COLUMNS]
    columns = ALL_COLUMNS + extra_cols
    enriched_col_idxs = {i + 1 for i, c in enumerate(columns) if c in OUTCOME_COLS}

    ws.append(columns)
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border    = BORDER

    for r_idx, row_data in enumerate(rows, start=2):
        ws.append([row_data.get(col, "") for col in columns])
        alt = r_idx % 2 == 0
        for c_idx, cell in enumerate(ws[r_idx], start=1):
            cell.font      = DATA_FONT
            cell.alignment = WRAP_ALIGN
            cell.border    = BORDER
            if c_idx in enriched_col_idxs:
                cell.fill = ENRICH_FILL
            elif alt:
                cell.fill = ALT_FILL

    for col_idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(col_name, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ---- Sheet 2: Score Summary (if scoring was run) -------------------------
    if score_result:
        ws2 = wb.create_sheet(title="Score Summary")

        SCORE_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
        SCORE_HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        SECTION_FILL       = PatternFill("solid", fgColor="2E86AB")
        SECTION_FONT       = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        LABEL_FONT         = Font(name="Arial", bold=True, size=10)
        VALUE_FONT         = Font(name="Arial", size=10)
        SCORE_FILL         = PatternFill("solid", fgColor="D4EDDA")  # light green
        MISSING_FILL       = PatternFill("solid", fgColor="FFF3CD")  # light yellow
        WRAP               = Alignment(wrap_text=True, vertical="top")
        CENTER             = Alignment(horizontal="center", vertical="center")
        THIN2              = Side(style="thin", color="CCCCCC")
        BDR                = Border(left=THIN2, right=THIN2, top=THIN2, bottom=THIN2)

        eps = score_result.get("endpoints", {})
        molecule = score_result.get("molecule", "")
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 22
        ws2.column_dimensions["C"].width = 22
        ws2.column_dimensions["D"].width = 18
        ws2.column_dimensions["E"].width = 18
        ws2.column_dimensions["F"].width = 55

        row = 1

        # Title
        ws2.merge_cells(f"A{row}:F{row}")
        title_cell = ws2[f"A{row}"]
        title_cell.value      = f"Dimension III – Clinical Efficacy Score: {molecule}"
        title_cell.font       = SCORE_HEADER_FONT
        title_cell.fill       = SCORE_HEADER_FILL
        title_cell.alignment  = CENTER
        row += 1

        # Overall score
        ws2.merge_cells(f"A{row}:F{row}")
        score_cell = ws2[f"A{row}"]
        ws_val = score_result.get("weighted_score", 0)
        score_cell.value     = f"Weighted Score: {ws_val:.3f} / 5.0   |   {score_result.get('data_coverage','')}"
        score_cell.font      = Font(name="Arial", bold=True, size=12, color="1F4E79")
        score_cell.fill      = SCORE_FILL
        score_cell.alignment = CENTER
        row += 1
        row += 1  # spacer

        # Endpoint breakdown header
        for col_letter, header in zip(
            ["A", "B", "C", "D", "E", "F"],
            ["Endpoint (Weight)", "Raw Value (%)", "Adj. Value (%)", "Phase Used", "Score (1-5)", "Trial Used / Reason"],
        ):
            c = ws2[f"{col_letter}{row}"]
            c.value     = header
            c.font      = SECTION_FONT
            c.fill      = SECTION_FILL
            c.alignment = CENTER
            c.border    = BDR
        row += 1

        # One row per endpoint
        ep_labels = {
            "weight_loss": "Weight Loss (40%)",
            "hba1c":       "HbA1c Reduction (40%)",
            "mash":        "MASH Resolution (10%)",
            "alt":         "ALT Reduction (10%)",
        }
        for ep_key, ep_label in ep_labels.items():
            ep = eps.get(ep_key, {})
            td = ep.get("trial_details", {})
            has_score = ep.get("score") is not None
            fill = SCORE_FILL if has_score else MISSING_FILL

            values = [
                ep_label,
                f"{ep.get('raw_value', 'N/A')}" if has_score else "N/A",
                f"{ep.get('best_value', 'N/A')}" if has_score else "N/A",
                f"Phase {ep.get('phase_used', 'N/A')}" if has_score else "N/A",
                str(ep.get("score", "N/A")),
                f"{td.get('trial_id','N/A')} | {td.get('dosage','N/A')} | {ep.get('reason','N/A')}",
            ]
            for col_letter, val in zip(["A", "B", "C", "D", "E", "F"], values):
                c = ws2[f"{col_letter}{row}"]
                c.value     = val
                c.font      = VALUE_FONT
                c.fill      = fill
                c.alignment = WRAP
                c.border    = BDR
            row += 1

        row += 1  # spacer

        # Score breakdown text block
        ws2.merge_cells(f"A{row}:F{row}")
        hdr = ws2[f"A{row}"]
        hdr.value     = "Score Breakdown"
        hdr.font      = SECTION_FONT
        hdr.fill      = SECTION_FILL
        hdr.alignment = CENTER
        row += 1

        for line in score_result.get("score_breakdown", "").split("\n"):
            ws2.merge_cells(f"A{row}:F{row}")
            c = ws2[f"A{row}"]
            c.value     = line
            c.font      = Font(name="Courier New", size=9)
            c.alignment = Alignment(wrap_text=False, vertical="top")
            row += 1

        row += 1  # spacer

        # Narrative rationale
        if score_rationale:
            ws2.merge_cells(f"A{row}:F{row}")
            hdr2 = ws2[f"A{row}"]
            hdr2.value     = "Clinical Narrative Rationale"
            hdr2.font      = SECTION_FONT
            hdr2.fill      = SECTION_FILL
            hdr2.alignment = CENTER
            row += 1

            ws2.merge_cells(f"A{row}:F{row + 14}")  # reserve ~15 rows
            rat_cell = ws2[f"A{row}"]
            rat_cell.value     = score_rationale
            rat_cell.font      = Font(name="Arial", size=10)
            rat_cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws2.row_dimensions[row].height = 300

    wb.save(path)
    print(f"\n[OUTPUT] Wrote {len(rows)} row(s) -> {path}", file=sys.stderr)
    if score_result:
        print(
            f"[OUTPUT] Score Summary sheet added  "
            f"(weighted score: {score_result.get('weighted_score','?')} / 5.0)",
            file=sys.stderr,
        )


# ==============================================================================
# SECTION 9 – CLI
# ==============================================================================

def _resolve_input_excel(molecule: str, explicit: Optional[str]) -> str:
    """Locate the Excel file to process."""
    if explicit:
        if not os.path.exists(explicit):
            sys.exit(f"ERROR: File not found: {explicit}")
        return explicit

    candidate = f"{molecule.lower().replace(' ', '_')}_trials.xlsx"
    if os.path.exists(candidate):
        return candidate

    candidates = [
        f for f in os.listdir(".")
        if f.endswith("_trials.xlsx") or f.endswith("_trials_enriched.xlsx")
    ]
    if len(candidates) == 1:
        print(f"  i Auto-discovered input file: {candidates[0]}", file=sys.stderr)
        return candidates[0]

    sys.exit(
        f"ERROR: Could not find input Excel file.\n"
        f"  Expected: {candidate}\n"
        f"  Or use:   --excel <path>"
    )


def main() -> int:
    ap = argparse.ArgumentParser(
        description=(
            "Enrich clinical trial Excel (from main1.py) with efficacy outcomes + dosage, "
            "then compute the Dimension III clinical efficacy score."
        )
    )
    ap.add_argument("molecule", help="Molecule / drug name (e.g. Cagrisema)")
    ap.add_argument("--excel",       default=None,  help="Input Excel file (default: <molecule>_trials.xlsx)")
    ap.add_argument("--out",         default=None,  help="Output Excel file (default: <molecule>_trials_enriched.xlsx)")
    ap.add_argument("--workers",     type=int, default=DEFAULT_WORKERS,
                    help=f"Max concurrent Gemini calls (default: {DEFAULT_WORKERS})")
    ap.add_argument("--run-main1",   action="store_true",
                    help="Run main1.py first to generate the input Excel, then enrich")
    ap.add_argument("--max-records", type=int, default=None,
                    help="Passed to main1.py if --run-main1 is set")
    ap.add_argument("--top-n",       type=int, default=None,
                    help="Passed to main1.py if --run-main1 is set")
    ap.add_argument("--no-score",    action="store_true",
                    help="Skip the Dimension III scoring step")
    args = ap.parse_args()

    molecule = args.molecule.strip()
    slug     = molecule.lower().replace(" ", "_")
    out_path = args.out or f"{slug}_trials_enriched.xlsx"

    print(f"\n{'='*60}", file=sys.stderr)
    print(f"  FETCHER  -  {molecule}", file=sys.stderr)
    print(f"{'='*60}\n", file=sys.stderr)

    _check_api_key_early()

    # Optionally run main1.py first
    if args.run_main1:
        import subprocess
        cmd = [sys.executable, "main1.py", molecule, "--no-enrich"]
        if args.max_records:
            cmd += ["--max-records", str(args.max_records)]
        if args.top_n:
            cmd += ["--top-n", str(args.top_n)]
        print(f"> Running: {' '.join(cmd)}\n", file=sys.stderr)
        result = subprocess.run(cmd)
        if result.returncode != 0:
            sys.exit("ERROR: main1.py failed.")

    # Load input Excel
    excel_path = _resolve_input_excel(molecule, args.excel)
    rows = _read_excel(excel_path)
    if not rows:
        print("No rows found in input Excel. Nothing to enrich.", file=sys.stderr)
        return 1

    # Step 1: Async enrichment
    t0 = time.time()
    enriched_rows = asyncio.run(enrich_all(molecule, rows, max_workers=args.workers))
    enrich_time = time.time() - t0
    print(f"[ENRICH] Time: {enrich_time:.1f}s", file=sys.stderr)

    # Step 2: Dimension III scoring
    score_result: Optional[Dict[str, Any]] = None
    score_rationale: Optional[str] = None

    if not args.no_score:
        print(f"\n[SCORE] Computing Dimension III Clinical Efficacy Score...", file=sys.stderr)
        score_result = compute_clinical_efficacy_score(molecule, enriched_rows)

        print(f"\n[SCORE] Results for {molecule}:", file=sys.stderr)
        print(f"  Weighted Score : {score_result['weighted_score']} / 5.0", file=sys.stderr)
        print(f"  Coverage       : {score_result['data_coverage']}", file=sys.stderr)
        print(f"  Breakdown:\n{score_result['score_breakdown']}", file=sys.stderr)

        score_rationale = generate_score_rationale(molecule, score_result)
    else:
        print("\n[SCORE] --no-score set, skipping scoring step.", file=sys.stderr)

    # Write output
    _write_excel(enriched_rows, out_path, score_result=score_result, score_rationale=score_rationale)

    total_time = time.time() - t0
    print(
        f"\nDone!  Output: {out_path}\n"
        f"  Rows          : {len(enriched_rows)}\n"
        f"  Enrich time   : {enrich_time:.1f}s\n"
        f"  Total time    : {total_time:.1f}s\n",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())